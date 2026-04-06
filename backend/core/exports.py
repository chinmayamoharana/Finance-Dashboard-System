import csv
import io
from typing import Any

from django.http import HttpResponse
from django.utils import timezone


class MissingExportDependencyError(RuntimeError):
    pass


SUPPORTED_EXPORT_FORMATS = {'csv', 'xlsx', 'pdf'}


def _timestamp_suffix() -> str:
    return timezone.localtime().strftime('%Y%m%d-%H%M%S')


def _stringify(value: Any) -> str:
    if value is None:
        return ''
    return str(value)


def _content_disposition(filename: str) -> str:
    return f'attachment; filename="{filename}"'


def _csv_response(filename_prefix: str, columns: list[dict[str, Any]], rows: list[dict[str, Any]]):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = _content_disposition(
        f'{filename_prefix}-{_timestamp_suffix()}.csv'
    )
    writer = csv.writer(response)
    writer.writerow([column['label'] for column in columns])
    for row in rows:
        writer.writerow([_stringify(row.get(column['key'])) for column in columns])
    return response


def _xlsx_response(
    filename_prefix: str,
    sheet_name: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise MissingExportDependencyError(
            'Excel export requires the "openpyxl" package to be installed.'
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    header_fill = PatternFill(fill_type='solid', fgColor='0F766E')
    header_font = Font(color='FFFFFF', bold=True)

    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=column['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            value = row.get(column['key'])
            cell = sheet.cell(row=row_index, column=column_index, value=_stringify(value))
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for index, column in enumerate(columns, start=1):
        max_length = len(column['label'])
        for row in rows:
            max_length = max(max_length, len(_stringify(row.get(column['key']))))
        sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 32)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = _content_disposition(
        f'{filename_prefix}-{_timestamp_suffix()}.xlsx'
    )
    return response


def _pdf_response(
    filename_prefix: str,
    title: str,
    subtitle: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise MissingExportDependencyError(
            'PDF export requires the "reportlab" package to be installed.'
        ) from exc

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()

    table_data = [[column['label'] for column in columns]]
    for row in rows:
        table_data.append([_stringify(row.get(column['key'])) for column in columns])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
        )
    )

    story = [
        Paragraph(title, styles['Title']),
        Spacer(1, 8),
        Paragraph(subtitle, styles['BodyText']),
        Spacer(1, 16),
        table,
    ]

    document.build(story)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = _content_disposition(
        f'{filename_prefix}-{_timestamp_suffix()}.pdf'
    )
    return response


def export_records_response(queryset, export_format: str):
    columns = [
        {'key': 'id', 'label': 'ID'},
        {'key': 'transaction_date', 'label': 'Transaction Date'},
        {'key': 'record_type', 'label': 'Type'},
        {'key': 'category', 'label': 'Category'},
        {'key': 'notes', 'label': 'Notes'},
        {'key': 'amount', 'label': 'Amount'},
        {'key': 'created_by', 'label': 'Created By'},
        {'key': 'updated_by', 'label': 'Updated By'},
        {'key': 'created_at', 'label': 'Created At'},
        {'key': 'updated_at', 'label': 'Updated At'},
    ]
    rows = [
        {
            'id': record.id,
            'transaction_date': record.transaction_date.isoformat(),
            'record_type': record.get_record_type_display(),
            'category': record.category,
            'notes': record.notes,
            'amount': f'{record.amount:.2f}',
            'created_by': record.created_by.name,
            'updated_by': record.updated_by.name if record.updated_by else '',
            'created_at': timezone.localtime(record.created_at).strftime('%Y-%m-%d %H:%M'),
            'updated_at': timezone.localtime(record.updated_at).strftime('%Y-%m-%d %H:%M'),
        }
        for record in queryset
    ]
    filename_prefix = 'financial-records-report'
    subtitle = f'Generated on {timezone.localtime().strftime("%d %b %Y %I:%M %p")}'

    if export_format == 'csv':
        return _csv_response(filename_prefix, columns, rows)
    if export_format == 'xlsx':
        return _xlsx_response(filename_prefix, 'Records', columns, rows)
    return _pdf_response(filename_prefix, 'Financial Records Report', subtitle, columns, rows)


def export_users_response(queryset, export_format: str):
    columns = [
        {'key': 'id', 'label': 'ID'},
        {'key': 'username', 'label': 'Username'},
        {'key': 'name', 'label': 'Name'},
        {'key': 'email', 'label': 'Email'},
        {'key': 'role', 'label': 'Role'},
        {'key': 'status', 'label': 'Status'},
        {'key': 'date_joined', 'label': 'Joined'},
        {'key': 'created_at', 'label': 'Created At'},
        {'key': 'updated_at', 'label': 'Updated At'},
    ]
    rows = [
        {
            'id': user.id,
            'username': user.username,
            'name': user.name,
            'email': user.email,
            'role': user.get_role_display(),
            'status': 'Active' if user.is_active else 'Inactive',
            'date_joined': timezone.localtime(user.date_joined).strftime('%Y-%m-%d %H:%M'),
            'created_at': timezone.localtime(user.created_at).strftime('%Y-%m-%d %H:%M'),
            'updated_at': timezone.localtime(user.updated_at).strftime('%Y-%m-%d %H:%M'),
        }
        for user in queryset
    ]
    filename_prefix = 'users-report'
    subtitle = f'Generated on {timezone.localtime().strftime("%d %b %Y %I:%M %p")}'

    if export_format == 'csv':
        return _csv_response(filename_prefix, columns, rows)
    if export_format == 'xlsx':
        return _xlsx_response(filename_prefix, 'Users', columns, rows)
    return _pdf_response(filename_prefix, 'Users Report', subtitle, columns, rows)
